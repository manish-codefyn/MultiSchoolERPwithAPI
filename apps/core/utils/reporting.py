import csv
import io
import pandas as pd
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

class ReportGenerator:
    """Utility class to generate reports in various formats."""

    @staticmethod
    def _get_value(obj, field):
        """Helper to get value from object or dictionary."""
        if isinstance(obj, dict):
            val = obj.get(field, "")
        else:
            val = getattr(obj, field, "")
            if callable(val):
                val = val()
        return val if val is not None else ""

    @staticmethod
    def generate_csv(data, columns, filename=None):
        """Generate a CSV response from data (queryset or list)."""
        if not filename:
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        elif not filename.endswith('.csv'):
            filename += '.csv'

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(columns)

        for obj in data:
            row = [ReportGenerator._get_value(obj, col) for col in columns]
            writer.writerow(row)

        return response

    @staticmethod
    def generate_excel(data, columns, filename=None, sheet_name="Report"):
        """Generate an Excel response from data."""
        if not filename:
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        elif not filename.endswith('.xlsx'):
            filename += '.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

        # Write data
        for row_num, obj in enumerate(data, 2):
            for col_num, col in enumerate(columns, 1):
                val = ReportGenerator._get_value(obj, col)
                ws.cell(row=row_num, column=col_num, value=str(val))

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @staticmethod
    def generate_pdf(data, columns, filename=None, title="Report Summary"):
        """Generate a PDF response using ReportLab."""
        if not filename:
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        elif not filename.endswith('.pdf'):
            filename += '.pdf'

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        title_style = styles['Title']
        normal_style = styles['Normal']

        # Add Title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 12))

        # Add Timestamp
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
        elements.append(Spacer(1, 24))

        # Table Data
        table_data = [columns]
        for obj in data:
            row = [str(ReportGenerator._get_value(obj, col)) for col in columns]
            table_data.append(row)

        # Create Table
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(t)
        doc.build(elements)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.getvalue())
        buffer.close()
        return response
